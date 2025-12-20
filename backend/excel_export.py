"""
Excel Export module for ServerScout
Generates professional Excel reports with multiple sheets
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime


# Export directory
EXPORT_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'exports')


def ensure_export_directory():
    """Ensure the exports directory exists"""
    if not os.path.exists(EXPORT_PATH):
        os.makedirs(EXPORT_PATH)


def get_cell_style():
    """Return standard cell styles"""
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_alignment = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    return {
        'header_font': header_font,
        'header_fill': header_fill,
        'header_alignment': header_alignment,
        'cell_alignment': cell_alignment,
        'border': thin_border
    }


def auto_adjust_column_width(worksheet):
    """Auto-adjust column widths based on content"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set width with some padding, max 50 characters
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = max(adjusted_width, 10)


def create_summary_sheet(workbook, stats):
    """Create the Summary sheet"""
    ws = workbook.create_sheet("Summary", 0)
    styles = get_cell_style()
    
    # Title
    ws['A1'] = 'SERVER INVENTORY SUMMARY'
    ws['A1'].font = Font(bold=True, size=16, color='2E7D32')
    ws.merge_cells('A1:B1')
    
    # Stats table
    summary_data = [
        ['Metric', 'Value'],
        ['Total Servers', stats.get('total', 0)],
        ['Online', stats.get('online', 0)],
        ['Offline', stats.get('offline', 0)],
        ['Not Scanned', stats.get('not_scanned', 0)],
        ['Last Scan', stats.get('last_scan', 'N/A') or 'Never'],
        ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles['border']
            
            if row_idx == 3:  # Header row
                cell.font = styles['header_font']
                cell.fill = styles['header_fill']
                cell.alignment = styles['header_alignment']
            else:
                cell.alignment = styles['cell_alignment']
    
    auto_adjust_column_width(ws)


def create_inventory_sheet(workbook, servers):
    """Create the Inventory sheet with all server details"""
    ws = workbook.create_sheet("Inventory", 1)
    styles = get_cell_style()
    
    # Define columns
    columns = [
        'Hostname', 'IP', 'Domain', 'OS Type', 'Brand', 'Model', 'Serial',
        'Motherboard', 'CPU Count', 'CPU Cores', 'CPU Logical', 'CPU Model',
        'RAM Physical', 'RAM Logical (MB)', 'Disk Info', 'Network Primary',
        'OS Version', 'Service Pack', 'Status', 'Last Scan'
    ]
    
    # Write header
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']
    
    # Write data
    for row_idx, server in enumerate(servers, start=2):
        data = [
            server.get('hostname', ''),
            server.get('ip', ''),
            server.get('domain', ''),
            server.get('os_type', ''),
            server.get('brand', ''),
            server.get('model', ''),
            server.get('serial', ''),
            server.get('motherboard', ''),
            server.get('cpu_count', ''),
            server.get('cpu_cores', ''),
            server.get('cpu_logical_processors', ''),
            server.get('cpu_model', ''),
            server.get('ram_physical', ''),
            server.get('ram_logical', ''),
            server.get('disk_info', ''),
            server.get('network_primary', ''),
            server.get('os_version', ''),
            server.get('service_pack', ''),
            server.get('status', ''),
            server.get('last_scan', '')
        ]
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = styles['cell_alignment']
            cell.border = styles['border']
            
            # Color coding for status
            if col_idx == 19:  # Status column
                if value == 'Online':
                    cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                elif value == 'Offline':
                    cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    auto_adjust_column_width(ws)


def create_warnings_sheet(workbook, servers):
    """Create the Warnings sheet for servers with issues"""
    ws = workbook.create_sheet("Warnings", 2)
    styles = get_cell_style()
    
    # Define columns
    columns = ['Hostname', 'IP', 'Warning', 'Details']
    
    # Write header
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid')
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']
    
    # Collect warnings
    warnings = []
    
    for server in servers:
        # Offline servers
        if server.get('status') == 'Offline':
            warnings.append({
                'hostname': server.get('hostname', server.get('ip', 'Unknown')),
                'ip': server.get('ip', ''),
                'warning': 'Server Offline',
                'details': 'Server is not responding'
            })
        
        # Not scanned servers
        if server.get('status') == 'Not Scanned':
            warnings.append({
                'hostname': server.get('hostname', server.get('ip', 'Unknown')),
                'ip': server.get('ip', ''),
                'warning': 'Not Scanned',
                'details': 'Server has never been scanned'
            })
    
    # Write warnings
    if warnings:
        for row_idx, warning in enumerate(warnings, start=2):
            data = [
                warning['hostname'],
                warning['ip'],
                warning['warning'],
                warning['details']
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = styles['cell_alignment']
                cell.border = styles['border']
                
                # Color coding for warning type
                if 'Offline' in warning['warning']:
                    cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                elif 'High' in warning['warning']:
                    cell.fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    else:
        ws.cell(row=2, column=1, value='No warnings found').alignment = styles['cell_alignment']
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    auto_adjust_column_width(ws)


def generate_excel_report(servers, stats):
    """
    Generate a complete Excel report
    
    Args:
        servers: list of server dictionaries
        stats: dictionary with server statistics
    
    Returns:
        str: path to generated Excel file
    """
    ensure_export_directory()
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Create sheets
    create_summary_sheet(wb, stats)
    create_inventory_sheet(wb, servers)
    create_warnings_sheet(wb, servers)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'inventory_{timestamp}.xlsx'
    filepath = os.path.join(EXPORT_PATH, filename)
    
    # Save workbook
    wb.save(filepath)
    
    return filepath


def get_export_filename():
    """Generate a timestamped filename for export"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f'inventory_{timestamp}.xlsx'


# ==================== PROJECT-BASED EXPORT FUNCTIONS ====================

def create_project_inventory_sheet(workbook, sheet_name, servers, sheet_index=None):
    """Create an inventory sheet for a specific project"""
    # Truncate sheet name to 31 characters (Excel limit)
    safe_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
    
    if sheet_index is not None:
        ws = workbook.create_sheet(safe_name, sheet_index)
    else:
        ws = workbook.create_sheet(safe_name)
    
    styles = get_cell_style()
    
    # Define columns
    columns = [
        'Hostname', 'IP', 'Domain', 'OS Type', 'Brand', 'Model', 'Serial',
        'Motherboard', 'CPU Count', 'CPU Cores', 'CPU Logical', 'CPU Model',
        'RAM Physical', 'RAM Logical (MB)', 'Disk Info', 'Network Primary',
        'OS Version', 'Service Pack', 'Status', 'Last Scan'
    ]
    
    # Write header
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']
    
    # Write data
    for row_idx, server in enumerate(servers, start=2):
        data = [
            server.get('hostname', ''),
            server.get('ip', ''),
            server.get('domain', ''),
            server.get('os_type', ''),
            server.get('brand', ''),
            server.get('model', ''),
            server.get('serial', ''),
            server.get('motherboard', ''),
            server.get('cpu_count', ''),
            server.get('cpu_cores', ''),
            server.get('cpu_logical_processors', ''),
            server.get('cpu_model', ''),
            server.get('ram_physical', ''),
            server.get('ram_logical', ''),
            server.get('disk_info', ''),
            server.get('network_primary', ''),
            server.get('os_version', ''),
            server.get('service_pack', ''),
            server.get('status', ''),
            server.get('last_scan', '')
        ]
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = styles['cell_alignment']
            cell.border = styles['border']
            
            # Color coding for status
            if col_idx == 19:  # Status column
                if value == 'Online':
                    cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                elif value == 'Offline':
                    cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    auto_adjust_column_width(ws)
    return ws


def create_project_summary_sheet(workbook, project_name, stats):
    """Create a summary sheet for a single project"""
    ws = workbook.create_sheet("Summary", 0)
    styles = get_cell_style()
    
    # Title
    ws['A1'] = f'PROJECT: {project_name}'
    ws['A1'].font = Font(bold=True, size=16, color='2E7D32')
    ws.merge_cells('A1:B1')
    
    # Stats table
    summary_data = [
        ['Metric', 'Value'],
        ['Total Servers', stats.get('total', 0)],
        ['Online', stats.get('online', 0)],
        ['Offline', stats.get('offline', 0)],
        ['Not Scanned', stats.get('not_scanned', 0)],
        ['Last Scan', stats.get('last_scan', 'N/A') or 'Never'],
        ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles['border']
            
            if row_idx == 3:  # Header row
                cell.font = styles['header_font']
                cell.fill = styles['header_fill']
                cell.alignment = styles['header_alignment']
            else:
                cell.alignment = styles['cell_alignment']
    
    auto_adjust_column_width(ws)


def create_all_projects_summary_sheet(workbook, projects_data, unassigned_stats):
    """Create a comparison summary sheet for all projects"""
    ws = workbook.create_sheet("Summary", 0)
    styles = get_cell_style()
    
    # Title
    ws['A1'] = 'ALL PROJECTS SUMMARY'
    ws['A1'].font = Font(bold=True, size=16, color='2E7D32')
    ws.merge_cells('A1:F1')
    
    # Report generation time
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(italic=True, size=10)
    
    # Headers
    columns = ['Project', 'Total', 'Online', 'Offline', 'Not Scanned', 'Last Scan']
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']
    
    # Project rows
    row_idx = 5
    total_all = {'total': 0, 'online': 0, 'offline': 0, 'not_scanned': 0}
    
    for project in projects_data:
        stats = project['stats']
        data = [
            project['name'],
            stats.get('total', 0),
            stats.get('online', 0),
            stats.get('offline', 0),
            stats.get('not_scanned', 0),
            stats.get('last_scan', 'Never') or 'Never'
        ]
        
        total_all['total'] += stats.get('total', 0)
        total_all['online'] += stats.get('online', 0)
        total_all['offline'] += stats.get('offline', 0)
        total_all['not_scanned'] += stats.get('not_scanned', 0)
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = styles['cell_alignment']
            cell.border = styles['border']
        
        row_idx += 1
    
    # Unassigned row (if there are unassigned servers)
    if unassigned_stats.get('total', 0) > 0:
        data = [
            '(Atanmamış)',
            unassigned_stats.get('total', 0),
            unassigned_stats.get('online', 0),
            unassigned_stats.get('offline', 0),
            unassigned_stats.get('not_scanned', 0),
            unassigned_stats.get('last_scan', 'Never') or 'Never'
        ]
        
        total_all['total'] += unassigned_stats.get('total', 0)
        total_all['online'] += unassigned_stats.get('online', 0)
        total_all['offline'] += unassigned_stats.get('offline', 0)
        total_all['not_scanned'] += unassigned_stats.get('not_scanned', 0)
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = styles['cell_alignment']
            cell.border = styles['border']
            cell.fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
        
        row_idx += 1
    
    # Total row
    row_idx += 1  # Empty row
    total_data = ['TOPLAM', total_all['total'], total_all['online'], total_all['offline'], total_all['not_scanned'], '']
    for col_idx, value in enumerate(total_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.alignment = styles['cell_alignment']
        cell.border = styles['border']
        cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    
    auto_adjust_column_width(ws)


def create_all_projects_warnings_sheet(workbook, projects_data, unassigned_servers):
    """Create a combined warnings sheet for all projects"""
    ws = workbook.create_sheet("Warnings", 1)
    styles = get_cell_style()
    
    # Define columns (with Project column)
    columns = ['Project', 'Hostname', 'IP', 'Warning', 'Details']
    
    # Write header
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid')
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']
    
    # Collect warnings from all projects
    warnings = []
    
    for project in projects_data:
        project_name = project['name']
        for server in project['servers']:
            if server.get('status') == 'Offline':
                warnings.append({
                    'project': project_name,
                    'hostname': server.get('hostname', server.get('ip', 'Unknown')),
                    'ip': server.get('ip', ''),
                    'warning': 'Server Offline',
                    'details': 'Server is not responding'
                })
            
            if server.get('status') == 'Not Scanned':
                warnings.append({
                    'project': project_name,
                    'hostname': server.get('hostname', server.get('ip', 'Unknown')),
                    'ip': server.get('ip', ''),
                    'warning': 'Not Scanned',
                    'details': 'Server has never been scanned'
                })
    
    # Add unassigned server warnings
    for server in unassigned_servers:
        if server.get('status') == 'Offline':
            warnings.append({
                'project': '(Atanmamış)',
                'hostname': server.get('hostname', server.get('ip', 'Unknown')),
                'ip': server.get('ip', ''),
                'warning': 'Server Offline',
                'details': 'Server is not responding'
            })
        
        if server.get('status') == 'Not Scanned':
            warnings.append({
                'project': '(Atanmamış)',
                'hostname': server.get('hostname', server.get('ip', 'Unknown')),
                'ip': server.get('ip', ''),
                'warning': 'Not Scanned',
                'details': 'Server has never been scanned'
            })
    
    # Write warnings
    if warnings:
        for row_idx, warning in enumerate(warnings, start=2):
            data = [
                warning['project'],
                warning['hostname'],
                warning['ip'],
                warning['warning'],
                warning['details']
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = styles['cell_alignment']
                cell.border = styles['border']
                
                # Color coding for warning type
                if 'Offline' in warning['warning']:
                    cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                elif 'Not Scanned' in warning['warning']:
                    cell.fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    else:
        ws.cell(row=2, column=1, value='No warnings found').alignment = styles['cell_alignment']
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    auto_adjust_column_width(ws)


def generate_project_excel_report(project_name, servers, stats):
    """
    Generate an Excel report for a single project
    
    Args:
        project_name: name of the project
        servers: list of server dictionaries for this project
        stats: dictionary with server statistics for this project
    
    Returns:
        str: path to generated Excel file
    """
    ensure_export_directory()
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Create sheets
    create_project_summary_sheet(wb, project_name, stats)
    create_project_inventory_sheet(wb, "Inventory", servers, 1)
    create_warnings_sheet(wb, servers)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_project_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).strip()
    filename = f'{safe_project_name}_{timestamp}.xlsx'
    filepath = os.path.join(EXPORT_PATH, filename)
    
    # Save workbook
    wb.save(filepath)
    
    return filepath


def generate_all_projects_excel_report(projects_data, unassigned_servers, unassigned_stats):
    """
    Generate an Excel report with all projects (each project as a separate sheet)
    
    Args:
        projects_data: list of dicts with 'name', 'servers', 'stats' for each project
        unassigned_servers: list of servers without a project
        unassigned_stats: stats for unassigned servers
    
    Returns:
        str: path to generated Excel file
    """
    ensure_export_directory()
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Create Summary sheet (comparison of all projects)
    create_all_projects_summary_sheet(wb, projects_data, unassigned_stats)
    
    # Create Warnings sheet (all warnings combined)
    create_all_projects_warnings_sheet(wb, projects_data, unassigned_servers)
    
    # Create a sheet for each project
    sheet_index = 2
    for project in projects_data:
        create_project_inventory_sheet(wb, project['name'], project['servers'], sheet_index)
        sheet_index += 1
    
    # Create sheet for unassigned servers (if any)
    if unassigned_servers:
        create_project_inventory_sheet(wb, "Atanmamış", unassigned_servers, sheet_index)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'all_projects_{timestamp}.xlsx'
    filepath = os.path.join(EXPORT_PATH, filename)
    
    # Save workbook
    wb.save(filepath)
    
    return filepath


# ==================== COMPARISON REPORT ====================

def generate_comparison_report(data, compare_type='scan-hpsm'):
    """Generate a 2-way comparison report Excel file"""
    ensure_export_directory()
    
    wb = Workbook()
    styles = get_cell_style()
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Determine source names
    if compare_type == 'scan-zabbix':
        source1 = 'Scan Report'
        source2 = 'Zabbix'
        count1 = data.get('scan_count', 0)
        count2 = data.get('zabbix_count', 0)
        missing_in_2 = data.get('missing_in_zabbix', [])
        missing_in_1 = data.get('missing_in_scan', [])
    elif compare_type == 'hpsm-zabbix':
        source1 = 'HPSM'
        source2 = 'Zabbix'
        count1 = data.get('hpsm_count', 0)
        count2 = data.get('zabbix_count', 0)
        missing_in_2 = data.get('missing_in_zabbix', [])
        missing_in_1 = data.get('missing_in_hpsm', [])
    else:  # scan-hpsm
        source1 = 'Scan Report'
        source2 = 'HPSM'
        count1 = data.get('scan_count', 0)
        count2 = data.get('hpsm_count', 0)
        missing_in_2 = data.get('missing_in_hpsm', [])
        missing_in_1 = data.get('missing_in_scan', [])
    
    matching = data.get('matching', [])
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = 'COMPARISON REPORT SUMMARY'
    ws_summary['A1'].font = Font(bold=True, size=16, color='2E7D32')
    ws_summary.merge_cells('A1:B1')
    
    summary_data = [
        ['Metric', 'Value'],
        [f'{source1} Servers', count1],
        [f'{source2} Servers', count2],
        ['Matching Servers', len(matching)],
        [f'Missing in {source2}', len(missing_in_2)],
        [f'Missing in {source1}', len(missing_in_1)],
        ['Report Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles['border']
            if row_idx == 3:
                cell.font = styles['header_font']
                cell.fill = styles['header_fill']
                cell.alignment = styles['header_alignment']
            else:
                cell.alignment = styles['cell_alignment']
    
    auto_adjust_column_width(ws_summary)
    
    # Missing in Source2 sheet
    ws_missing_2 = wb.create_sheet(f"Missing in {source2}", 1)
    ws_missing_2['A1'] = f'SERVERS MISSING IN {source2.upper()}'
    ws_missing_2['A1'].font = Font(bold=True, size=14, color='D32F2F')
    ws_missing_2.cell(row=2, column=1, value=f'These servers are in {source1} but not in {source2}').font = Font(italic=True, size=10)
    
    header_cell = ws_missing_2.cell(row=4, column=1, value='Hostname')
    header_cell.font = styles['header_font']
    header_cell.fill = PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid')
    header_cell.alignment = styles['header_alignment']
    header_cell.border = styles['border']
    
    if missing_in_2:
        for row_idx, hostname in enumerate(missing_in_2, start=5):
            cell = ws_missing_2.cell(row=row_idx, column=1, value=hostname)
            cell.alignment = styles['cell_alignment']
            cell.border = styles['border']
    else:
        ws_missing_2.cell(row=5, column=1, value='No missing records').font = Font(italic=True)
    
    auto_adjust_column_width(ws_missing_2)
    
    # Missing in Source1 sheet
    ws_missing_1 = wb.create_sheet(f"Missing in {source1}", 2)
    ws_missing_1['A1'] = f'SERVERS MISSING IN {source1.upper()}'
    ws_missing_1['A1'].font = Font(bold=True, size=14, color='F59E0B')
    ws_missing_1.cell(row=2, column=1, value=f'These servers are in {source2} but not in {source1}').font = Font(italic=True, size=10)
    
    header_cell = ws_missing_1.cell(row=4, column=1, value='Hostname')
    header_cell.font = styles['header_font']
    header_cell.fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
    header_cell.alignment = styles['header_alignment']
    header_cell.border = styles['border']
    
    if missing_in_1:
        for row_idx, hostname in enumerate(missing_in_1, start=5):
            cell = ws_missing_1.cell(row=row_idx, column=1, value=hostname)
            cell.alignment = styles['cell_alignment']
            cell.border = styles['border']
    else:
        ws_missing_1.cell(row=5, column=1, value='No missing records').font = Font(italic=True)
    
    auto_adjust_column_width(ws_missing_1)
    
    # Matching sheet
    ws_matching = wb.create_sheet("Matching", 3)
    ws_matching['A1'] = 'MATCHING SERVERS'
    ws_matching['A1'].font = Font(bold=True, size=14, color='22C55E')
    
    header_cell = ws_matching.cell(row=3, column=1, value='Hostname')
    header_cell.font = styles['header_font']
    header_cell.fill = PatternFill(start_color='22C55E', end_color='22C55E', fill_type='solid')
    header_cell.alignment = styles['header_alignment']
    header_cell.border = styles['border']
    
    if matching:
        for row_idx, hostname in enumerate(matching, start=4):
            cell = ws_matching.cell(row=row_idx, column=1, value=hostname)
            cell.alignment = styles['cell_alignment']
            cell.border = styles['border']
    else:
        ws_matching.cell(row=4, column=1, value='No matching records').font = Font(italic=True)
    
    auto_adjust_column_width(ws_matching)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'comparison_report_{timestamp}.xlsx'
    filepath = os.path.join(EXPORT_PATH, filename)
    
    wb.save(filepath)
    return filepath


def generate_full_comparison_report(data):
    """Generate a 3-way comparison report Excel file"""
    ensure_export_directory()
    
    wb = Workbook()
    styles = get_cell_style()
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = 'FULL 3-WAY COMPARISON REPORT'
    ws_summary['A1'].font = Font(bold=True, size=16, color='8B5CF6')
    ws_summary.merge_cells('A1:B1')
    
    in_all_three = data.get('in_all_three', [])
    in_scan_hpsm = data.get('in_scan_and_hpsm_only', [])
    in_scan_zabbix = data.get('in_scan_and_zabbix_only', [])
    in_hpsm_zabbix = data.get('in_hpsm_and_zabbix_only', [])
    only_scan = data.get('only_in_scan', [])
    only_hpsm = data.get('only_in_hpsm', [])
    only_zabbix = data.get('only_in_zabbix', [])
    
    summary_data = [
        ['Metric', 'Count'],
        ['Scan Report Total', data.get('scan_count', 0)],
        ['HPSM Report Total', data.get('hpsm_count', 0)],
        ['Zabbix Report Total', data.get('zabbix_count', 0)],
        ['', ''],
        ['VERIFIED (In All 3)', len(in_all_three)],
        ['In Scan + HPSM only', len(in_scan_hpsm)],
        ['In Scan + Zabbix only', len(in_scan_zabbix)],
        ['In HPSM + Zabbix only', len(in_hpsm_zabbix)],
        ['Only in Scan', len(only_scan)],
        ['Only in HPSM', len(only_hpsm)],
        ['Only in Zabbix', len(only_zabbix)],
        ['', ''],
        ['Report Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles['border']
            if row_idx == 3:
                cell.font = styles['header_font']
                cell.fill = styles['header_fill']
                cell.alignment = styles['header_alignment']
            elif row_idx == 8:  # Verified row
                cell.font = Font(bold=True, color='22C55E')
                cell.alignment = styles['cell_alignment']
            else:
                cell.alignment = styles['cell_alignment']
    
    auto_adjust_column_width(ws_summary)
    
    # Helper function to create list sheet
    def create_list_sheet(ws_name, title, color, items, description, index):
        ws = wb.create_sheet(ws_name, index)
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14, color=color)
        ws.cell(row=2, column=1, value=description).font = Font(italic=True, size=10)
        
        header_cell = ws.cell(row=4, column=1, value='Hostname')
        header_cell.font = styles['header_font']
        header_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        header_cell.alignment = styles['header_alignment']
        header_cell.border = styles['border']
        
        if items:
            for row_idx, hostname in enumerate(items, start=5):
                cell = ws.cell(row=row_idx, column=1, value=hostname)
                cell.alignment = styles['cell_alignment']
                cell.border = styles['border']
        else:
            ws.cell(row=5, column=1, value='No records').font = Font(italic=True)
        
        auto_adjust_column_width(ws)
    
    # Create sheets for each category
    create_list_sheet("Verified (All 3)", "VERIFIED - IN ALL 3 SYSTEMS", '22C55E', in_all_three, 
                      "These servers are confirmed in Scan, HPSM, and Zabbix", 1)
    
    create_list_sheet("Scan+HPSM Only", "IN SCAN AND HPSM ONLY", 'F59E0B', in_scan_hpsm,
                      "Missing from Zabbix monitoring", 2)
    
    create_list_sheet("Scan+Zabbix Only", "IN SCAN AND ZABBIX ONLY", 'F59E0B', in_scan_zabbix,
                      "Missing from HPSM inventory", 3)
    
    create_list_sheet("HPSM+Zabbix Only", "IN HPSM AND ZABBIX ONLY", 'F59E0B', in_hpsm_zabbix,
                      "Missing from Scan report - may need scanning", 4)
    
    create_list_sheet("Only in Scan", "ONLY IN SCAN REPORT", 'D32F2F', only_scan,
                      "Review needed - not in HPSM or Zabbix", 5)
    
    create_list_sheet("Only in HPSM", "ONLY IN HPSM", 'D32F2F', only_hpsm,
                      "Review needed - not in Scan or Zabbix", 6)
    
    create_list_sheet("Only in Zabbix", "ONLY IN ZABBIX", 'D32F2F', only_zabbix,
                      "Review needed - not in Scan or HPSM", 7)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'full_comparison_report_{timestamp}.xlsx'
    filepath = os.path.join(EXPORT_PATH, filename)
    
    wb.save(filepath)
    return filepath

